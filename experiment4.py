from __future__ import annotations

import math
import os
from dataclasses import dataclass
from typing import List, Tuple

from openpyxl import Workbook, load_workbook


@dataclass
class Student:
    name: str
    age: int
    chinese: float
    math: float
    english: float

    @property
    def total(self) -> float:
        return self.chinese + self.math + self.english

    @property
    def average(self) -> float:
        return self.total / 3


class Vector3:
    def __init__(self, x: float, y: float, z: float):
        self.x = x
        self.y = y
        self.z = z

    def __add__(self, other: "Vector3") -> "Vector3":
        return Vector3(self.x + other.x, self.y + other.y, self.z + other.z)

    def __sub__(self, other: "Vector3") -> "Vector3":
        return Vector3(self.x - other.x, self.y - other.y, self.z - other.z)

    def __repr__(self) -> str:
        return f"Vector3({self.x}, {self.y}, {self.z})"


def ensure_excel(path: str) -> None:
    if os.path.exists(path):
        os.remove(path)
    data = [
        ("张甜甜", 21, 97, 99, 95),
        ("章程", 23, 99, 99, 81),
        ("赵小明", 19, 68, 97, 75),
        ("王华", 20, 75, 80, 79),
        ("李梅", 21, 70, 72, 70),
        ("陈强", 20, 72, 74, 70),
        ("张敏", 22, 65, 70, 68),
        ("刘芳", 20, 70, 75, 80),
        ("吴磊", 21, 80, 78, 76),
        ("周洁", 20, 66, 83, 77),
    ]
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Age", "Chinese_Score", "Math_Score", "English_Score"])
    for row in data:
        ws.append(row)
    wb.save(path)


def load_students(path: str) -> List[Student]:
    wb = load_workbook(path)
    ws = wb.active
    students: List[Student] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            continue
        name, age, chinese, math, english = row
        students.append(Student(str(name), int(age), float(chinese), float(math), float(english)))
    return students


def string_ops() -> Tuple[str, str, str, str, str]:
    s = "2018 Amazon Jeff Bezos 1120"
    part1 = s.replace("2018 ", "", 1)
    digits = "".join(ch for ch in s if ch.isdigit())
    bracketed = s.replace("2018", "【2018】", 1).replace("1120", "【1120】", 1)
    no_space = s.replace(" ", "")
    doubled = []
    for token in s.split():
        if token.isdigit():
            doubled.append(str(int(token) * 2))
        else:
            doubled.append(token)
    doubled_s = " ".join(doubled)
    return part1, digits, bracketed, no_space, doubled_s


def factorial_sum(n: int) -> int:
    total = 0
    current = 1
    for i in range(1, n + 1):
        current *= i
        total += current
    return total


if __name__ == "__main__":
    excel_path = os.path.join(os.path.dirname(__file__), "学生成绩数据.xlsx")
    ensure_excel(excel_path)
    students = load_students(excel_path)
    avg_chinese = sum(s.chinese for s in students) / len(students)
    avg_math = sum(s.math for s in students) / len(students)
    avg_english = sum(s.english for s in students) / len(students)
    print(f"语文平均分: {avg_chinese:.2f}")
    print(f"数学平均分: {avg_math:.2f}")
    print(f"英语平均分: {avg_english:.2f}")
    top3 = sorted(students, key=lambda x: x.total, reverse=True)[:3]
    print("\n总成绩前三名学生信息:")
    for i, stu in enumerate(top3, 1):
        print(f"第{i}名:")
        print(f"姓名: {stu.name}, 年龄: {stu.age}, 语文: {int(stu.chinese)}, 数学: {int(stu.math)}, 英语: {int(stu.english)}")
        print(f"平均分: {stu.average:.2f}, 总成绩: {int(stu.total)}")
        if i != len(top3):
            print()

    v1 = Vector3(1, 2, 3)
    v2 = Vector3(4, 5, 6)
    print("v1+v2", v1 + v2)
    print("v1-v2", v1 - v2)

    part1, digits, bracketed, no_space, doubled = string_ops()
    print(part1)
    print(digits)
    print(bracketed)
    print(no_space)
    print(doubled)

    n = 6
    print("factorial sum", factorial_sum(n))

